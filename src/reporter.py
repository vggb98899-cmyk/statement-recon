"""
reporter.py — 对账结果输出模块

职责：
  将 matched + classified 的结果输出为：
    Sheet 1: 对账结果明细  — 每一笔订单的对账结果
    Sheet 2: 异常清单      — 差异 > 阈值的记录（标红）
    Sheet 3: 孤儿流水      — 支付侧有、平台侧找不到的记录

格式：
  Excel (.xlsx)，带条件颜色：
    🔴 红色 — 异常（差异 > 2元）
    🟡 黄色 — 手续费差异
    🟢 绿色 — 匹配成功
"""

import os
from datetime import datetime

import pandas as pd

# ============================================================
# 颜色定义（openpyxl 格式）
# ============================================================

# 异常红色
RED_FILL = "FFE0E0"       # 浅红背景
RED_FONT = "CC0000"       # 深红文字

# 手续费差异黄色
YELLOW_FILL = "FFF9C4"    # 浅黄背景
YELLOW_FONT = "8D6E00"    # 棕色文字

# 匹配绿色
GREEN_FILL = "E8F5E9"     # 浅绿背景
GREEN_FONT = "2E7D32"     # 深绿文字

# 表头样式
HEADER_FILL = "4472C4"    # 蓝色背景
HEADER_FONT = "FFFFFF"    # 白色文字


def _apply_style(worksheet, row_idx, classification):
    """
    根据分类结果，给行设置颜色。
    """
    from openpyxl.styles import Font, PatternFill

    if "数据缺失" in classification or "待查" in classification:
        fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
        font = Font(color=RED_FONT)
    elif "手续费" in classification:
        fill = PatternFill(start_color=YELLOW_FILL, end_color=YELLOW_FILL, fill_type="solid")
        font = Font(color=YELLOW_FONT)
    elif "匹配" in classification:
        fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
        font = Font(color=GREEN_FONT)
    else:
        return  # 不处理未知分类

    for cell in worksheet[row_idx]:
        cell.fill = fill
        cell.font = font


def _write_header(worksheet, headers):
    """
    写表头并设置样式。
    """
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(color=HEADER_FONT, bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font


def _auto_width(worksheet, headers, max_width=40):
    """
    自动调整列宽。
    """
    for col_idx, header in enumerate(headers, 1):
        worksheet.column_dimensions[chr(64 + col_idx)].width = min(len(str(header)) + 4, max_width)


def write_reconciliation_report(
    matched_df: pd.DataFrame,
    orphan_df: pd.DataFrame,
    anomaly_df: pd.DataFrame,
    output_dir: str = "output",
    threshold: float = 2.0,
) -> str:
    """
    输出对账结果 Excel 报告。

    参数：
      matched_df — classify_all() 的输出（含 classification 列）
      orphan_df  — matcher.match_orders() 的孤儿流水
      anomaly_df — classifier.get_anomalies() 的输出
      output_dir — 输出目录
      threshold  — 差异阈值（用于报告标题标注）

    返回：
      输出文件路径
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 生成文件名：对账结果_YYYYMMDD_HHMMSS.xlsx
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"对账结果_{timestamp}.xlsx")

    # 创建工作簿
    wb = Workbook()

    # ============================================================
    # Sheet 1: 对账结果明细
    # ============================================================
    ws1 = wb.active
    ws1.title = "对账结果"

    # 定义输出列
    detail_cols = [
        ("order_id", "订单号"),
        ("platform", "平台"),
        ("total", "平台应收"),
        ("platform_fee", "平台手续费"),
        ("expected_net", "预期净到账"),
        ("Settlement Amount", "支付到账"),
        ("Fee", "支付手续费"),
        ("difference", "差异金额"),
        ("classification", "差异归类"),
        ("subtag", "差异子标签"),
    ]

    headers = [label for _, label in detail_cols]
    _write_header(ws1, headers)

    for row_idx, (_, row) in enumerate(matched_df.iterrows(), 2):
        for col_idx, (col_name, _) in enumerate(detail_cols, 1):
            val = row.get(col_name, "")
            # 格式化金额保留两位小数
            if isinstance(val, float):
                val = round(val, 2)
            ws1.cell(row=row_idx, column=col_idx, value=val)

        # 根据分类设置行颜色
        _apply_style(ws1, row_idx, str(row.get("classification", "")))

    # ---- v1.1: 底部汇总统计 ----
    summary_row = len(matched_df) + 3  # 空一行再写
    total = len(matched_df)
    matched_count = len(matched_df[matched_df["classification"].str.contains("✅", na=False)])
    fee_diff = len(matched_df[matched_df["classification"].str.contains("手续费", na=False)])
    pending = len(matched_df[matched_df["classification"].str.contains("待查", na=False)])
    missing = len(matched_df[matched_df["classification"].str.contains("数据缺失", na=False)])

    orphan_count = len(orphan_df)
    ws1.cell(row=summary_row, column=1,
             value=f"总笔数: {total}  |  ✅ 匹配: {matched_count}  |  ⚠️ 手续费差异: {fee_diff}  |  ❌ 待查: {pending}  |  ❌ 数据缺失: {missing}  |  🕊️ 孤儿流水: {orphan_count}")

    _auto_width(ws1, headers)

    # ============================================================
    # Sheet 2: 异常清单
    # ============================================================
    ws2 = wb.create_sheet(title="异常清单")

    anomaly_cols = [
        ("order_id", "订单号"),
        ("platform", "平台"),
        ("total", "平台应收"),
        ("Settlement Amount", "支付到账"),
        ("difference", "差异金额"),
        ("classification", "差异归类"),
    ]

    headers2 = [label for _, label in anomaly_cols]
    _write_header(ws2, headers2)

    # 如果异常清单为空，写一行提示
    if anomaly_df.empty:
        ws2.cell(row=2, column=1, value="🎉 无异常记录，所有差异均在阈值范围内。")
    else:
        for row_idx, (_, row) in enumerate(anomaly_df.iterrows(), 2):
            for col_idx, (col_name, _) in enumerate(anomaly_cols, 1):
                val = row.get(col_name, "")
                if isinstance(val, float):
                    val = round(val, 2)
                ws2.cell(row=row_idx, column=col_idx, value=val)

            # 异常清单全部标红
            _apply_style(ws2, row_idx, "❌ 待查")

    _auto_width(ws2, headers2)

    # ============================================================
    # Sheet 3: 孤儿流水
    # ============================================================
    ws3 = wb.create_sheet(title="孤儿流水")

    # v1.1: 孤儿流水的建议处理动作（固定模板）
    ORPHAN_SUGGESTION = (
        "待查明：该订单号在三个平台均无记录，"
        "建议联系支付渠道（PingPong）获取该笔流水的原始交易凭证，"
        "核实是否为其他店铺误入账。"
    )

    orphan_cols = [
        ("Transaction ID", "支付流水号"),
        ("order_id", "关联订单号"),
        ("Settlement Amount", "到账金额"),
        ("Fee", "手续费"),
        ("Currency", "币种"),
        ("_suggestion", "建议处理动作"),
    ]

    headers3 = [label for _, label in orphan_cols]
    _write_header(ws3, headers3)

    if orphan_df.empty:
        ws3.cell(row=2, column=1, value="🎉 无孤儿流水，支付侧记录均有对应的平台订单。")
    else:
        for row_idx, (_, row) in enumerate(orphan_df.iterrows(), 2):
            for col_idx, (col_name, _) in enumerate(orphan_cols, 1):
                if col_name == "_suggestion":
                    val = ORPHAN_SUGGESTION
                else:
                    val = row.get(col_name, "")
                    if isinstance(val, float):
                        val = round(val, 2)
                ws3.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws3, headers3)

    # ============================================================
    # 保存
    # ============================================================
    wb.save(output_path)
    return output_path


def print_summary(matched_df: pd.DataFrame):
    """
    在控制台打印汇总统计。
    让用户一眼知道对账结果的总览。
    """
    total = len(matched_df)
    matched_count = len(matched_df[matched_df["classification"].str.contains("✅", na=False)])
    fee_diff = len(matched_df[matched_df["classification"].str.contains("手续费", na=False)])
    pending = len(matched_df[matched_df["classification"].str.contains("待查", na=False)])
    missing = len(matched_df[matched_df["classification"].str.contains("数据缺失", na=False)])

    max_diff = matched_df["difference"].max() if not matched_df.empty else 0
    min_diff = matched_df["difference"].min() if not matched_df.empty else 0

    print("=" * 50)
    print("📊 对账汇总")
    print("=" * 50)
    print(f"  总订单数:      {total}")
    print(f"  ✅ 匹配:      {matched_count}")
    print(f"  ⚠️ 手续费差异: {fee_diff}")
    print(f"  ❌ 待查:      {pending}")
    print(f"  ❌ 数据缺失:   {missing}")
    print(f"  最大差异:     {max_diff:.2f} 元")
    print(f"  最小差异:     {min_diff:.2f} 元")
    print("=" * 50)
